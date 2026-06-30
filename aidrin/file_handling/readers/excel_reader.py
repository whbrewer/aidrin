import os

import openpyxl
import pandas as pd

from aidrin.file_handling.readers.base_reader import BaseFileReader

_OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm"}


class excelReader(BaseFileReader):
    def read(self):
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext not in _OPENPYXL_EXTENSIONS:
            return pd.read_excel(self.file_path)

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active

        num_header_rows = self._detect_header_rows(ws)
        if num_header_rows == 1:
            return pd.read_excel(self.file_path)

        header_matrix = self._get_header_matrix(ws, num_header_rows)
        flat_columns = self._flatten_header(header_matrix)
        return pd.read_excel(
            self.file_path,
            skiprows=num_header_rows,
            header=None,
            names=flat_columns,
        )

    def _detect_header_rows(self, ws):
        merged_ranges = list(ws.merged_cells.ranges)
        if not merged_ranges:
            return 1

        # Only consider merged ranges that start in the top two rows.
        top_merges = [r for r in merged_ranges if r.min_row <= 2]
        if not top_merges:
            return 1

        max_merge_row = max(r.max_row for r in top_merges)
        if max_merge_row > 1:
            # A vertical/block merge already tells us the header depth.
            return max_merge_row

        # All merges are horizontal (within row 1 only). Check whether row 2
        # also contains only string labels — if so it is the leaf-name row.
        if self._row_looks_like_header(ws, 2):
            return 2

        return 1

    def _row_looks_like_header(self, ws, row_num):
        for cell in ws[row_num]:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if cell.value is not None and not isinstance(cell.value, str):
                return False
        return True

    def _get_header_matrix(self, ws, num_header_rows):
        max_col = ws.max_column
        matrix = [[None] * max_col for _ in range(num_header_rows)]

        for row_idx in range(1, num_header_rows + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    matrix[row_idx - 1][col_idx - 1] = cell.value

        for merge_range in ws.merged_cells.ranges:
            if merge_range.min_row > num_header_rows:
                continue
            top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
            for r in range(merge_range.min_row, min(merge_range.max_row, num_header_rows) + 1):
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    matrix[r - 1][c - 1] = top_left_value

        return matrix

    def _flatten_header(self, matrix):
        num_cols = len(matrix[0]) if matrix else 0
        columns = []
        for col_idx in range(num_cols):
            seen = {}
            parts = []
            for row_idx in range(len(matrix)):
                val = matrix[row_idx][col_idx]
                if val is not None:
                    label = str(val).strip()
                    if label and label not in seen:
                        parts.append(label)
                        seen[label] = True
            columns.append(" | ".join(parts) if parts else f"Column_{col_idx + 1}")
        return columns

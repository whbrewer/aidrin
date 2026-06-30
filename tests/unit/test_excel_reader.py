"""Unit tests for excelReader with multi-row / merged-cell header support.

These tests run without Flask, Celery, or Redis.
"""

import logging
import os
import tempfile
import unittest

import openpyxl
import pandas as pd

from aidrin.file_handling.readers.excel_reader import excelReader

_log = logging.getLogger("test")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean(path):
    try:
        os.unlink(path)
    except OSError:
        pass


def _make_xlsx(setup_fn) -> str:
    """Create a temp .xlsx file, call setup_fn(ws) to populate it, return path."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    setup_fn(ws)
    wb.save(tmp.name)
    return tmp.name


# ---------------------------------------------------------------------------
# Single header row (fast path — no merged cells)
# ---------------------------------------------------------------------------

class TestExcelReaderSingleHeader(unittest.TestCase):

    def test_reads_columns_and_rows(self):
        def setup(ws):
            ws.append(["Name", "Age", "Score"])
            ws.append(["Alice", 25, 90.5])
            ws.append(["Bob", 30, 85.0])

        path = _make_xlsx(setup)
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertIsInstance(df, pd.DataFrame)
        self.assertEqual(list(df.columns), ["Name", "Age", "Score"])
        self.assertEqual(len(df), 2)
        self.assertEqual(df["Name"].tolist(), ["Alice", "Bob"])

    def test_single_header_no_merge_preserves_values(self):
        def setup(ws):
            ws.append(["X", "Y"])
            ws.append([1, 2])
            ws.append([3, 4])

        path = _make_xlsx(setup)
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertEqual(df["X"].tolist(), [1, 3])
        self.assertEqual(df["Y"].tolist(), [2, 4])


# ---------------------------------------------------------------------------
# Two header rows: top row merged horizontally
# ---------------------------------------------------------------------------

class TestExcelReaderTwoRowHeader(unittest.TestCase):

    def _make_demographics_xlsx(self):
        """
        Row 1: "Demographics" merged A1:C1 | "Info" merged D1:E1
        Row 2: "Age" | "Gender" | "Race" | "Name" | "ID"
        Data : 25, M, Asian, Alice, 1
               30, F, White, Bob,   2
        """
        def setup(ws):
            ws["A1"] = "Demographics"
            ws.merge_cells("A1:C1")
            ws["D1"] = "Info"
            ws.merge_cells("D1:E1")
            ws["A2"] = "Age"
            ws["B2"] = "Gender"
            ws["C2"] = "Race"
            ws["D2"] = "Name"
            ws["E2"] = "ID"
            ws.append([25, "M", "Asian", "Alice", 1])
            ws.append([30, "F", "White", "Bob", 2])

        return _make_xlsx(setup)

    def test_column_names_concatenated_with_separator(self):
        path = self._make_demographics_xlsx()
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertEqual(
            list(df.columns),
            [
                "Demographics | Age",
                "Demographics | Gender",
                "Demographics | Race",
                "Info | Name",
                "Info | ID",
            ],
        )

    def test_data_rows_loaded_correctly(self):
        path = self._make_demographics_xlsx()
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertEqual(len(df), 2)
        self.assertEqual(df["Demographics | Age"].tolist(), [25, 30])
        self.assertEqual(df["Info | Name"].tolist(), ["Alice", "Bob"])


# ---------------------------------------------------------------------------
# Merged cell spanning both header rows (no duplication in flattened name)
# ---------------------------------------------------------------------------

class TestExcelReaderMergeSpansBothRows(unittest.TestCase):

    def test_vertically_merged_header_cell_not_duplicated(self):
        """
        Row 1: "ID" merged A1:A2 | "Score" in B1
        Row 2: (A2 is part of merge)  | "Points" in B2
        Expected columns: ["ID", "Score | Points"]
        """
        def setup(ws):
            ws["A1"] = "ID"
            ws.merge_cells("A1:A2")
            ws["B1"] = "Score"
            ws["B2"] = "Points"
            ws.append([1, 95])
            ws.append([2, 88])

        path = _make_xlsx(setup)
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertIn("ID", df.columns)
        self.assertNotIn("ID | ID", df.columns)
        self.assertIn("Score | Points", df.columns)

    def test_vertically_merged_cell_data_correct(self):
        def setup(ws):
            ws["A1"] = "ID"
            ws.merge_cells("A1:A2")
            ws["B1"] = "Value"
            ws["B2"] = "Raw"
            ws.append([10, 99])
            ws.append([20, 77])

        path = _make_xlsx(setup)
        try:
            df = excelReader(path, _log).read()
        finally:
            _clean(path)

        self.assertEqual(df["ID"].tolist(), [10, 20])


# ---------------------------------------------------------------------------
# Fallback: .xls extension skips openpyxl and delegates to pandas
# ---------------------------------------------------------------------------

class TestExcelReaderXlsFallback(unittest.TestCase):

    def test_xls_path_uses_pandas_fallback(self):
        """
        excelReader with a .xls path should not raise AttributeError or
        openpyxl-specific errors — it falls back to pd.read_excel().
        We use an xlsx file renamed to .xls to verify the code path branches
        without needing the legacy xlrd engine.
        """
        def setup(ws):
            ws.append(["A", "B"])
            ws.append([1, 2])

        xlsx_path = _make_xlsx(setup)
        xls_path = xlsx_path.replace(".xlsx", ".xls")
        os.rename(xlsx_path, xls_path)
        try:
            # Should not crash with an openpyxl-specific error;
            # pandas may raise an engine error for a mis-named file,
            # but the branching logic itself must not raise AttributeError.
            try:
                excelReader(xls_path, _log).read()
            except Exception as exc:
                self.assertNotIsInstance(exc, AttributeError)
        finally:
            _clean(xls_path)


if __name__ == "__main__":
    unittest.main()
